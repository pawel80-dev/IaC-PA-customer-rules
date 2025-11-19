from openpyxl import load_workbook
import logging

# logger will return the source module name
logger = logging.getLogger(__name__)
# display logging info level
logging.basicConfig(level=logging.INFO)


# More validations can be added here
def validate_template(file_path, sheet_name=None) -> None:
    excel_workbook = load_workbook(filename=file_path, data_only=True)
    excel_sheet = excel_workbook[sheet_name]
    temp_valid = []

    for row in excel_sheet.iter_rows(min_row=1, max_col=1, max_row=excel_sheet.max_row):
        for cell in row:
            temp_valid.append(cell.value)
    
    if ("Group_address_objects" not in temp_valid) or ("Policies" not in temp_valid):
        exit("The template file is not valid. Please fix the template structure.")
    else:
        logger.info("Template structure is valid.")


def service_port_strip(service: str) -> list:
    # logger.info(f"Service:\n{service}")
    serv_list = []
    if service == "application-default":
        serv_list.append(service)
        return serv_list
    if service == "any":
        serv_list.append(service)
        return serv_list
    else:
        split_lines = service.splitlines()
        for line in split_lines:
            serv_dict = {}
            serv_dict["name"] = line
            prot_port = line.split("_")
            serv_dict["protocol"] = prot_port[0]
            serv_dict["port"] = prot_port[1]
            serv_list.append(serv_dict)

        return serv_list


def create_config(file_path, sheet_name=None) -> None:
    excel_workbook = load_workbook(filename=file_path, data_only=True)
    excel_sheet = excel_workbook[sheet_name]
    cfg_addr_groups = []
    cfg_services = []
    cfg_policies = []
    all_configs = {}

    # Start from the 1st row to set the config_flag
    # row is a list of cells in that row
    for row in excel_sheet.iter_rows(min_row=1, max_col=excel_sheet.max_column, max_row=excel_sheet.max_row):
        # thanks to the flag system we scan the config only once
        if row[0].value == "Group_address_objects":
            config_flag = "addr_group"
        if row[0].value == "Policies":
            config_flag = "policy"
        if (row[0].value != "Group_address_objects" and row[0].value != "Group_name" 
            and config_flag == "addr_group" and row[0].value is not None
            and row[1].value is not None and row[2].value is not None):
            cfg_addr_group = {}
            cfg_addr_group["name"] = row[0].value
            cfg_addr_group["objects"] = [{row[1].value: row[2].value}]
            cfg_addr_groups.append(cfg_addr_group)
            # logger.info(f"ADDR_GROUP_DICT_FIRST_ROW:\n{row[0].value, row[1].value, row[2].value}")
        if row[0].value is None and row[1].value is not None and row[2].value is not None and config_flag == "addr_group":
            cfg_addr_group["objects"].extend([{row[1].value: row[2].value}])
            # logger.info(f"ADDR_GROUP_DICT_SECOND_ROW:\n{row[0].value, row[1].value, row[2].value}")

        # flag was changed, time for a second part of the config
        if (row[0].value != "Policies" and row[0].value != "Policy_name" 
            and config_flag == "policy" and row[0].value is not None):
            cfg_policy = {}
            cfg_policy["name"] = row[0].value
            cfg_policy["description"] = row[1].value
            cfg_policy["src_zone"] = row[2].value
            cfg_policy["src_addr"] = row[3].value
            cfg_policy["dst_zone"] = row[4].value
            cfg_policy["dst_addr"] = row[5].value
            cfg_policy["app"] = row[6].value
            cfg_policy["service"] = service_port_strip(row[7].value)
            cfg_policy["action"] = row[8].value
            # logger.info(f"POLICY_DICT:\n{cfg_policy}")
            cfg_policies.append(cfg_policy)
            # logger.info(f"CURRENT_POLICY_LIST:\n{cfg_policies}")
        if (row[0].value != "Policies" and row[0].value != "Policy_name" 
            and config_flag == "policy" and row[0].value is not None 
            and row[7].value != "application-default" and row[7].value != "any"):
            cfg_services.extend(service_port_strip(row[7].value))

    logger.info(f"Policies:\n{cfg_policies}")
    logger.info(f"Address groups:\n{cfg_addr_groups}")
    logger.info(f"Services:\n{cfg_services}")

    all_configs["policies"] = cfg_policies
    all_configs["addr_groups"] = cfg_addr_groups
    all_configs["services"] = cfg_services

    return all_configs


def main() -> None:
    validate_template("customer_rules/customer_A.xlsx", sheet_name="Rules")
    create_config("customer_rules/customer_A.xlsx", sheet_name="Rules")
    # read_excel_file("customer_rules/customer_A.xlsx", sheet_name="Rules")
    # ports = service_port_strip("tcp_7700\ntcp_7701\ntcp_7702")
    # ports = service_port_strip("udp_123-222")


if __name__ == "__main__":
    main()